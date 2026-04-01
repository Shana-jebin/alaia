from django.shortcuts import render, redirect, get_object_or_404,HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.db.models import Q,Sum
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect
from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from django.utils.text import slugify
import json
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from products.models import Product, ProductVariant, VariantImage, Brand, Category,Occasion,Coupon
from .forms import ProductForm, ProductVariantForm
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.contrib.admin.views.decorators import staff_member_required
from decimal import Decimal, InvalidOperation
from products.models import Review  
from django.utils.dateparse import parse_datetime
from django.utils import timezone as tz





User = get_user_model()

def is_admin(user):
    return user.is_authenticated and user.is_staff
@never_cache
@csrf_protect
def admin_login(request):

    if request.user.is_authenticated and request.user.is_staff:
        return redirect('admin_dashboard')

   
    if request.method == "POST":
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()

        user = authenticate(request, username=username, password=password)

        if user is None:
            messages.error(request, "Invalid credentials.")
        elif not user.is_staff:
            messages.error(request, "You don't have admin access.")
        elif not user.is_active:
            messages.error(request, "This account is blocked.")
        else:
            login(request, user)
            return redirect('admin_dashboard')

    return render(request, 'adminpanel/admin-login.html')
@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def admin_dashboard(request):
    from orders.models import Order, OrderItem
    from products.models import Product
    from django.db.models import Sum, Count
    from django.db.models.functions import TruncMonth, TruncDate
    from datetime import timedelta

    # ── User stats ──
    total_users   = User.objects.count()
    active_users  = User.objects.filter(is_active=True).count()
    blocked_users = User.objects.filter(is_active=False).count()
    today_start   = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_signups = User.objects.filter(date_joined__gte=today_start).count()

    # ── Order / revenue stats ──
    all_orders     = Order.objects.all()
    total_orders   = all_orders.count()
    total_revenue  = all_orders.filter(
        payment_status='paid'
    ).aggregate(t=Sum('total'))['t'] or 0

    # vs last month
    now            = timezone.now()
    this_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_month_end   = this_month_start - timedelta(seconds=1)
    last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    this_month_revenue = Order.objects.filter(
        payment_status='paid', created_at__gte=this_month_start
    ).aggregate(t=Sum('total'))['t'] or 0

    last_month_revenue = Order.objects.filter(
        payment_status='paid',
        created_at__gte=last_month_start,
        created_at__lte=last_month_end,
    ).aggregate(t=Sum('total'))['t'] or 0

    this_month_orders = Order.objects.filter(created_at__gte=this_month_start).count()
    last_month_orders = Order.objects.filter(
        created_at__gte=last_month_start,
        created_at__lte=last_month_end,
    ).count()

    def pct_change(current, previous):
        if previous == 0:
            return 100 if current > 0 else 0
        return round(((current - previous) / previous) * 100, 1)

    revenue_change = pct_change(float(this_month_revenue), float(last_month_revenue))
    orders_change  = pct_change(this_month_orders, last_month_orders)

    # ── Revenue chart — last 7 months ──
    seven_months_ago = now - timedelta(days=210)
    monthly_revenue  = (
        Order.objects.filter(
            payment_status='paid',
            created_at__gte=seven_months_ago,
        )
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total=Sum('total'))
        .order_by('month')
    )
    chart_labels  = [r['month'].strftime('%b') for r in monthly_revenue]
    chart_data    = [float(r['total']) for r in monthly_revenue]

    # ── Sales by category ──
    from orders.models import OrderItem
    category_sales = (
        OrderItem.objects
        .values('variant__product__category__name')
        .annotate(total=Sum('unit_price'))
        .order_by('-total')[:5]
    )
    cat_labels = [r['variant__product__category__name'] or 'Unknown' for r in category_sales]
    cat_data   = [float(r['total']) for r in category_sales]

    # ── Recent orders ──
    recent_orders = (
        Order.objects
        .select_related('user')
        .prefetch_related('items')
        .order_by('-created_at')[:5]
    )

    # ── Product stats ──
    total_products = Product.objects.filter(is_deleted=False).count()

    context = {
        'total_users':       total_users,
        'active_users':      active_users,
        'blocked_users':     blocked_users,
        'today_signups':     today_signups,
        'username':          request.user.get_full_name() or request.user.username,
        'total_orders':      total_orders,
        'total_revenue':     total_revenue,
        'revenue_change':    revenue_change,
        'orders_change':     orders_change,
        'this_month_revenue': this_month_revenue,
        'this_month_orders': this_month_orders,
        'total_products':    total_products,
        'chart_labels':      chart_labels,
        'chart_data':        chart_data,
        'cat_labels':        cat_labels,
        'cat_data':          cat_data,
        'recent_orders':     recent_orders,
    }
    return render(request, 'adminpanel/dashboard.html', context)

@never_cache
@login_required
@user_passes_test(is_admin)
def user_management(request):
    query = request.GET.get('q', '').strip()
    page_number = request.GET.get('page')

    users = User.objects.filter(is_superuser=False).order_by('-date_joined')


    if query:
        users = users.filter(
            Q(email__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(username__icontains=query)
        )

    paginator = Paginator(users, 10)
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'query': query,
    }

    return render(request, 'adminpanel/user_list.html', context)

@never_cache
@login_required
@user_passes_test(is_admin)
def toggle_block_user(request, user_id):
    if request.method == "POST":
        user = get_object_or_404(User, id=user_id)

        if user == request.user:
            messages.error(request, "You cannot block yourself.")
            return redirect('user_management')

        if user.is_active:
            user.is_active = False
            messages.success(request, f"{user.username} has been blocked.")
        else:
            user.is_active = True
            messages.success(request, f"{user.username} has been unblocked.")

        user.save()

    return redirect('user_management')




@login_required
@user_passes_test(is_admin)
def search_users(request):
    query = request.GET.get('q', '').strip()

    users = User.objects.filter(is_superuser=False)

    if query:
        users = users.filter(username__istartswith=query
)

    data = [
        {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "is_active": user.is_active,
            "date_joined": user.date_joined.strftime("%b %d, %Y")
        }
        for user in users
    ]

    return JsonResponse({"users": data})

def generate_unique_slug(name, instance=None):
    slug = slugify(name)
    unique_slug = slug
    counter = 1

    while True:
        slug_exists = Category.objects.filter(slug=unique_slug)

        if instance:
            slug_exists = slug_exists.exclude(id=instance.id)

        if not slug_exists.exists():
            break

        unique_slug = f"{slug}-{counter}"
        counter += 1

    return unique_slug


@login_required
@user_passes_test(is_admin)
def admin_category_list(request):
    query = request.GET.get('q', '')

    if request.method == "POST":
        name = request.POST.get('name', '').strip().title()


        if not name:
            messages.error(request, "Category name cannot be empty.")
            return redirect('admin_category_list')


        description = request.POST.get('description')
        offer_percentage = request.POST.get('offer_percentage')
        is_active = request.POST.get('is_active') == "True"

     
        existing_category = Category.objects.filter(
            name__iexact=name,
            is_deleted=True
        ).first()

        if existing_category:
            existing_category.is_deleted = False
            existing_category.description = description
            existing_category.offer_percentage = offer_percentage or None
            existing_category.is_active = is_active
            existing_category.slug = generate_unique_slug(name, instance=existing_category)
            existing_category.save()

            messages.success(request, "Deleted category restored successfully!")
            return redirect('admin_category_list')
        

       
        if Category.objects.filter(name__iexact=name, is_deleted=False).exists():
            messages.error(request, "Category already exists!")
            return redirect('admin_category_list')


        slug = generate_unique_slug(name)

        Category.objects.create(
            name=name,
            slug=slug,
            description=description,
            offer_percentage=offer_percentage if offer_percentage else None,
            is_active=is_active
        )

        messages.success(request, "Category added successfully!")
        return redirect('admin_category_list')

 
    categories = Category.objects.filter(is_deleted=False)

    if query:
        categories = categories.filter(name__icontains=query)

    categories = categories.order_by('-created_at')

    paginator = Paginator(categories, 5)
    page_number = request.GET.get('page', 1)

    try:
        page_obj = paginator.page(page_number)
    except:
        page_obj = paginator.page(paginator.num_pages)


    context = {
    'page_obj': page_obj,
    'query': query,
    'request': request
}

    return render(request, 'adminpanel/category-management.html', context)



@login_required
@user_passes_test(is_admin)
def admin_category_edit(request, category_id):
    category = get_object_or_404(Category, id=category_id)

    if request.method == "POST":
        name = request.POST.get('name', '').strip().title()

        # ❗ EMPTY NAME BLOCK
        if not name:
            messages.error(request, "Category name cannot be empty.")
            return redirect('admin_category_list')



        category.name = name
        category.slug = generate_unique_slug(name, instance=category)
        category.description = request.POST.get('description')
        category.offer_percentage = request.POST.get('offer_percentage') or None
        category.is_active = request.POST.get('is_active') == "True"

        category.save()

        messages.success(request, "Category updated successfully!")
        return redirect('admin_category_list')

    return redirect('admin_category_list')



@login_required
@user_passes_test(is_admin)
def admin_category_delete(request, category_id):
    category = get_object_or_404(Category, id=category_id)

    if request.method == "POST":
        category.is_deleted = True  
        category.save()

        messages.success(request, "Category deleted successfully!")

    return redirect('admin_category_list')






@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def product_list(request):
    
    search_query = request.GET.get('q', '')
    per_page = int(request.GET.get('per_page', 10))
    page_number = request.GET.get('page', 1)
    show_deleted = request.GET.get('show') == 'deleted'

    if per_page not in [5, 10, 25, 50]:
        per_page = 10

   

    if show_deleted:
        products = Product.all_objects.filter(is_deleted=True)
    else:
        products = Product.objects.all()
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(category__name__icontains=search_query) |
            Q(brand__name__icontains=search_query)
        )

    products = products.order_by('-created_at' if not show_deleted else '-deleted_at')


    all_active = Product.objects.filter(is_deleted=False)
    total_products = all_active.count()
    active_products = all_active.filter(is_active=True).count()
    inactive_products = all_active.filter(is_active=False).count()
    deleted_count = Product.objects.filter(is_deleted=True).count()

    paginator = Paginator(products, per_page)
    products_page = paginator.get_page(page_number)

    categories = Category.objects.filter(
    is_deleted=False,
    is_active=True
)
    brands = Brand.objects.all()

    context = {
        'products': products_page,
        'search_query': search_query,
        'per_page': per_page,
        'show_deleted': show_deleted,
        'categories': categories,
        'brands': brands,
        'total_products': total_products,
        'active_products': active_products,
        'inactive_products': inactive_products,
        'deleted_count': deleted_count,
        'occasions': Occasion.objects.all(),
    }
    return render(request, 'adminpanel/product_list.html', context)
@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def product_add(request):

    product_form = ProductForm(request.POST)

    if not product_form.is_valid():
        return JsonResponse({
            'success': False,
            'errors': product_form.errors
        }, status=400)

    product = product_form.save()

    variants_json = request.POST.get('variants')

    if not variants_json:
        return JsonResponse({
            'success': False,
            'message': 'Variants data is required.'
        }, status=400)

    try:
        variants_data = json.loads(variants_json)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid variants JSON'
        }, status=400)

    for index, variant_data in enumerate(variants_data, start=1):

        sizes = variant_data.get('sizes', [])

        image_key = f'variant_images_{index}'
        images = request.FILES.getlist(image_key)

        if len(images) < 3:
            return JsonResponse({
                'success': False,
                'message': 'Please upload at least 3 images.'
            }, status=400)

        first_variant = None

        for size in sizes:

            variant = ProductVariant.objects.create(
                product=product,
                color=variant_data.get('color'),
                size=size,
                price=float(variant_data.get('price') or 0),
                sales_price=float(variant_data.get('sales_price')) if variant_data.get('sales_price') else None,
                stock=int(variant_data.get('stock') or 0),
            )

            if not first_variant:
                first_variant = variant

        # save images only once
        if first_variant:
            for img in images:
                VariantImage.objects.create(
                    variant=first_variant,
                    image=img
                )

    return JsonResponse({
        'success': True,
        'product_id': product.id,
        'message': 'Product added successfully!'
    })
@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def product_edit(request, pk):

    product = get_object_or_404(Product, pk=pk, is_deleted=False)

    if request.method == 'POST':

        product_form = ProductForm(request.POST, instance=product)

        if not product_form.is_valid():
            return JsonResponse({
                'success': False,
                'errors': product_form.errors
            }, status=400)

        product = product_form.save()

        variants_json = request.POST.get('variants')

        if variants_json:
            try:
                variants_data = json.loads(variants_json)
            except json.JSONDecodeError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid variant data'
                }, status=400)

         
            old_images_by_color = {}
            for v in product.variants.all():
                imgs = list(v.images.all())
                if imgs:
                    old_images_by_color[v.color] = imgs

            product.variants.all().delete()

           

            for index, variant_data in enumerate(variants_data, start=1):

                sizes = variant_data.get('sizes', [])

                image_key = f'variant_images_{index}'
                images = request.FILES.getlist(image_key)

                first_variant = None

                for size in sizes:

                    variant = ProductVariant.objects.create(
                        product=product,
                        color=variant_data.get('color'),
                        size=size,
                        price=float(variant_data.get('price') or 0),
                        sales_price=float(variant_data.get('sales_price')) if variant_data.get('sales_price') else None,
                        stock=int(variant_data.get('stock') or 0),
                    )

                    if not first_variant:
                        first_variant = variant

                # save images
                if first_variant:
                    if images:
                        for img in images:
                            VariantImage.objects.create(
                                variant=first_variant,
                                image=img
                            )
                    else:
                        # Restore old images from saved map
                        color = variant_data.get('color')
                        old_imgs = old_images_by_color.get(color, [])
                        for old in old_imgs:
                            VariantImage.objects.create(
                                variant=first_variant,
                                image=old.image
                            )

        return JsonResponse({
            'success': True,
            'message': 'Product updated successfully!'
        })

    # ---------- GET PART (FOR EDIT FORM) ----------

    grouped = {}

    for v in product.variants.all():

        if v.color not in grouped:
            grouped[v.color] = {
                'id': v.id,
                'color': v.color,
                'sizes': [],
                'price': str(v.price),
                'sales_price': str(v.sales_price) if v.sales_price else '',
                'stock': v.stock,
                'images': []
            }

        grouped[v.color]['sizes'].append(v.size)

        for img in v.images.all():
            grouped[v.color]['images'].append({
                'id': img.id,
                'url': img.image.url
            })

    return JsonResponse({
        'id': product.id,
        'name': product.name,
        'description': product.description,
        'category': product.category_id,
        'brand': product.brand_id,
        'occasions': list(product.occasions.values_list('id', flat=True)),
        'is_active': product.is_active,
        'is_featured': product.is_featured,
        'variants': list(grouped.values())
    })
@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def product_soft_delete(request, pk):
    product = get_object_or_404(Product, pk=pk, is_deleted=False)
    product.soft_delete()
    return JsonResponse({'success': True, 'message': f'"{product.name}" has been deactivated (soft deleted).'})

@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def restore_product(request, pk):
    try:
        product = Product.all_objects.get(pk=pk)  
        product.restore()

        return JsonResponse({
            "success": True,
            "message": "Product restored successfully!"
        })

    except Product.DoesNotExist:
        return JsonResponse({
            "success": False,
            "message": "Product not found"
        })
    
@require_POST
def product_toggle_status(request, pk):
    product = get_object_or_404(Product, pk=pk, is_deleted=False)
    product.is_active = not product.is_active
    product.save()
    status = 'Active' if product.is_active else 'Inactive'
    return JsonResponse({'success': True, 'is_active': product.is_active, 'message': f'Product is now {status}.'})


def product_upload_image(request):
    if request.method == 'POST' and request.FILES.get('image'):
        image = request.FILES['image']
        variant_id = request.POST.get('variant_id')
        
        if variant_id:
            variant = get_object_or_404(ProductVariant, pk=variant_id)
            variant_image = VariantImage.objects.create(variant=variant, image=image)
            return JsonResponse({'success': True, 'image_id': variant_image.id, 'image_url': variant_image.image.url})
        
        return JsonResponse({'success': False, 'message': 'variant_id is required'}, status=400)
    
    return JsonResponse({'success': False, 'message': 'No image provided'}, status=400)


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def brand_list(request):

    query = request.GET.get('q','').strip()

    active_brands = Brand.objects.filter(is_active=True).order_by('name')
    inactive_brands = Brand.objects.filter(is_active=False).order_by('name')


    if query:
        active_brands = active_brands.filter(name__icontains=query)
        inactive_brands = inactive_brands.filter(name__icontains=query)


    paginator = Paginator(active_brands, 5)  
    page_number = request.GET.get('page')
    brands = paginator.get_page(page_number)

    return render(request,'adminpanel/brand-management.html',{
        'brands': brands,
        'inactive_brands': inactive_brands,
        'query': query,
    })


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def brand_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        rating = request.POST.get('rating')
        logo = request.FILES.get('logo')

        if not name:
            messages.error(request, "Brand name is required")
            return redirect('brand_create')

        if Brand.objects.filter(name__iexact=name).exists():
            messages.error(request, "Brand already exists")
            return redirect('brand_create')

        if rating:
            try:
                rating = float(rating)
                if rating < 0 or rating > 5:
                    messages.error(request, "Rating must be between 0 and 5")
                    return redirect('brand_create')
            except:
                messages.error(request, "Invalid rating value")
                return redirect('brand_create')
        else:
            rating = 0.0

        Brand.objects.create(
            name=name,
            rating=rating,
            logo=logo
        )

        messages.success(request, "Brand added successfully")
        return redirect('brand_list')


    return render(request, 'adminpanel/brand-management.html')


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def brand_edit(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name')
        rating = request.POST.get('rating')
        logo = request.FILES.get('logo')

   
        if not name:
            messages.error(request, "Brand name is required")
            return redirect('brand_edit', pk=pk)

        if Brand.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, "Brand already exists")
            return redirect('brand_edit', pk=pk)

        if rating:
            try:
                rating = float(rating)
                if rating < 0 or rating > 5:
                    messages.error(request, "Rating must be between 0 and 5")
                    return redirect('brand_edit', pk=pk)
            except:
                messages.error(request, "Invalid rating value")
                return redirect('brand_edit', pk=pk)
        else:
            rating = 0.0

        brand.name = name
        brand.rating = rating

        if logo:
            brand.logo = logo

        brand.save()

        messages.success(request, "Brand updated successfully")
        return redirect('brand_list')

    return render(request, 'adminpanel/brand-edit.html', {
        'brand': brand
    })

@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def brand_delete(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    brand.is_active = False
    brand.save()
    return redirect('brand_list')

@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def brand_restore(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    brand.is_active = True
    brand.save()
    return redirect('brand_list')



# ADD THESE IMPORTS AND VIEWS TO YOUR adminpanel/views.py

from orders.models import Order, OrderItem
from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncDate


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def admin_order_list(request):
    orders = Order.objects.select_related('user').prefetch_related('items')

    q = request.GET.get('q', '').strip()
    if q:
        orders = orders.filter(
            Q(order_id__icontains=q) |
            Q(user__email__icontains=q) |
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(full_name__icontains=q) |
            Q(items__product_name__icontains=q)
        ).distinct()

    status_filter = request.GET.get('status', '')
    if status_filter:
        orders = orders.filter(status=status_filter)

    date_from = request.GET.get('date_from', '')
    if date_from:
        try:
            from datetime import datetime
            orders = orders.filter(created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date())
        except ValueError:
            pass

    date_to = request.GET.get('date_to', '')
    if date_to:
        try:
            from datetime import datetime
            orders = orders.filter(created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date())
        except ValueError:
            pass

    sort = request.GET.get('sort', '-created_at')
    allowed_sorts = ['created_at', '-created_at', 'total', '-total', 'status']
    if sort in allowed_sorts:
        orders = orders.order_by(sort)
    else:
        orders = orders.order_by('-created_at')

    paginator = Paginator(orders, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    # Stats
    all_orders = Order.objects.all()
    stats = {
        'total':     all_orders.count(),
        'pending':   all_orders.filter(status='pending').count(),
        'shipped':   all_orders.filter(status='shipped').count(),
        'delivered': all_orders.filter(status='delivered').count(),
        'cancelled': all_orders.filter(status='cancelled').count(),
        'revenue':   all_orders.filter(status='delivered').aggregate(t=Sum('total'))['t'] or 0,
    }

    return render(request, 'adminpanel/order-list.html', {
        'page_obj': page_obj,
        'q': q,
        'status_filter': status_filter,
        'date_from': date_from,
        'date_to': date_to,
        'sort': sort,
        'stats': stats,
        'status_choices': Order.STATUS_CHOICES,
    })


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def admin_order_detail(request, order_id):
    order = get_object_or_404(Order, order_id=order_id)
    return render(request, 'adminpanel/order-detail.html', {
        'order': order,
        'status_choices': Order.STATUS_CHOICES,
    })


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def admin_order_status(request, order_id):
    order = get_object_or_404(Order, order_id=order_id)
    try:
        data = json.loads(request.body)
        new_status = data.get('status', '').strip()
    except Exception:
        new_status = request.POST.get('status', '').strip()

    valid_statuses = [s[0] for s in Order.STATUS_CHOICES]
    if new_status not in valid_statuses:
        return JsonResponse({'error': 'Invalid status.'}, status=400)

    
    if new_status == 'cancelled' and order.status not in ('cancelled', 'returned'):
        for item in order.items.filter(status='active'):
            if item.variant:
                item.variant.stock += item.quantity
                item.variant.save(update_fields=['stock'])
            item.status = 'cancelled'
            item.save()


    refund_msg = ''
    if new_status == 'returned' and order.status == 'return_requested':
        refund_amount = order.refund_amount()
        for item in order.items.filter(status='return_requested'):
            if item.variant:
                item.variant.stock += item.quantity
                item.variant.save(update_fields=['stock'])
            item.status = 'returned'
            item.save()
        if refund_amount > 0:
            wallet, _ = Wallet.objects.get_or_create(user=order.user)
            wallet.credit(
                refund_amount,
                description=f"Refund for returned order {order.order_id}",
                order=order,
            )
            refund_msg = f' ₹{refund_amount:.0f} refunded to customer wallet.'

    order.status = new_status
    order.save(update_fields=['status', 'updated_at'])

    return JsonResponse({
        'success': True,
        'message': f'Order status updated to {order.get_status_display()}.{refund_msg}',
        'new_status': new_status,
        'new_status_display': order.get_status_display(),
    })


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def admin_inventory(request):
    

    variants = ProductVariant.objects.filter(
    is_deleted=False,
    product__is_deleted=False,
).select_related('product', 'product__category', 'product__brand').order_by('stock')

    q = request.GET.get('q', '').strip()

    if q:
        variants = variants.filter(
            Q(product__name__istartswith=q) |
            Q(product__brand__name__istartswith=q) |
            Q(product__category__name__istartswith=q) |
            Q(color__istartswith=q) |
            Q(size__istartswith=q)
        ).distinct()

    stock_filter = request.GET.get('stock', '')
    if stock_filter == 'out':
        variants = variants.filter(stock=0)
    elif stock_filter == 'low':
        variants = variants.filter(stock__gt=0, stock__lte=10)
    elif stock_filter == 'ok':
        variants = variants.filter(stock__gt=10)

    paginator = Paginator(variants, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    stats = {
    'total':   ProductVariant.objects.filter(is_deleted=False, product__is_deleted=False).count(),
    'out':     ProductVariant.objects.filter(is_deleted=False, product__is_deleted=False, stock=0).count(),
    'low':     ProductVariant.objects.filter(is_deleted=False, product__is_deleted=False, stock__gt=0, stock__lte=10).count(),
    'healthy': ProductVariant.objects.filter(is_deleted=False, product__is_deleted=False, stock__gt=10).count(),
}

    return render(request, 'adminpanel/inventory.html', {
        'page_obj': page_obj,
        'q': q,
        'stock_filter': stock_filter,
        'stats': stats,
    })


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def admin_update_stock(request, variant_id):
    from products.models import ProductVariant
    variant = get_object_or_404(ProductVariant, id=variant_id)
    try:
        data = json.loads(request.body)
        new_stock = int(data.get('stock', 0))
        if new_stock < 0:
            return JsonResponse({'error': 'Stock cannot be negative.'}, status=400)
    except (ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({'error': 'Invalid stock value.'}, status=400)

    variant.stock = new_stock
    variant.save(update_fields=['stock'])
    return JsonResponse({
        'success': True,
        'stock': variant.stock,
        'message': f'Stock updated to {variant.stock}.',
    })




 
def _validate(data, pk=None):
    errors = []
 
    code = data.get('code', '').strip().upper()
    if not code:
        errors.append('Coupon code is required.')
    elif len(code) < 3:
        errors.append('Code must be at least 3 characters.')
    else:
        qs = Coupon.objects.filter(code__iexact=code)
        if pk:
            qs = qs.exclude(pk=pk)
        if qs.exists():
            errors.append(f'Code "{code}" already exists.')
 
    dtype = data.get('discount_type', '')
    if dtype not in ('percent', 'flat'):
        errors.append('Please select a discount type.')
 
    dval = data.get('discount_value', '').strip()
    if not dval:
        errors.append('Discount value is required.')
    else:
        try:
            v = Decimal(dval)
            if v <= 0:
                errors.append('Discount value must be greater than 0.')
            if dtype == 'percent' and v > 100:
                errors.append('Percentage discount cannot exceed 100%.')
        except InvalidOperation:
            errors.append('Enter a valid number for discount value.')
 
    vf = data.get('valid_from', '').strip()
    vt = data.get('valid_to',   '').strip()
    if not vf:
        errors.append('"Valid From" date is required.')
    if not vt:
        errors.append('"Valid To" date is required.')
    if vf and vt and vf >= vt:
        errors.append('"Valid To" must be after "Valid From".')
 
    raw_ul = data.get('usage_limit', '').strip()
    if raw_ul:
        try:
            if int(raw_ul) < 1:
                errors.append('Usage limit must be at least 1.')
        except ValueError:
            errors.append('Usage limit must be a whole number.')
 
    return errors
 
 
def _save_coupon(coupon, data):
    coupon.code           = data['code'].strip().upper()
    coupon.description    = data.get('description', '').strip()
    coupon.discount_type  = data['discount_type']
    coupon.discount_value = Decimal(data['discount_value'])
    coupon.min_order      = Decimal(data.get('min_order') or '0')
    coupon.max_discount   = Decimal(data['max_discount']) if data.get('max_discount', '').strip() else None
    vf = parse_datetime(data['valid_from'])
    vt = parse_datetime(data['valid_to'])

    if vf and tz.is_naive(vf):
        vf = tz.make_aware(vf)
    if vt and tz.is_naive(vt):
        vt = tz.make_aware(vt)

    coupon.valid_from = vf
    coupon.valid_to   = vt
    coupon.usage_limit    = int(data['usage_limit']) if data.get('usage_limit', '').strip() else None
    coupon.per_user_limit = int(data.get('per_user_limit') or 1)
    coupon.is_active      = data.get('is_active')  == 'on'
    coupon.is_one_time    = data.get('is_one_time') == 'on'
    coupon.save()
    return coupon
 

@staff_member_required
def coupon_list(request):
    qs = Coupon.objects.order_by('-created_at')
    q  = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(code__icontains=q) | Q(description__icontains=q))
 
    paginator = Paginator(qs, 15)
    page_obj  = paginator.get_page(request.GET.get('page'))
 
    return render(request, 'adminpanel/coupons.html', {
        'page_obj': page_obj,
        'q': q,
    })
 

@staff_member_required
@require_POST
def coupon_create(request):
    errors = _validate(request.POST)
    if errors:
        return JsonResponse({'success': False, 'errors': errors})
 
    coupon = _save_coupon(Coupon(), request.POST)
    return JsonResponse({'success': True, 'message': f'Coupon "{coupon.code}" created.'})
 

@staff_member_required
@require_POST
def coupon_edit(request, pk):
    coupon = get_object_or_404(Coupon, pk=pk)
    errors = _validate(request.POST, pk=pk)
    if errors:
        return JsonResponse({'success': False, 'errors': errors})
 
    coupon = _save_coupon(coupon, request.POST)
    return JsonResponse({'success': True, 'message': f'Coupon "{coupon.code}" updated.'})
 

@staff_member_required
@require_POST
def coupon_toggle(request, pk):
    coupon           = get_object_or_404(Coupon, pk=pk)
    coupon.is_active = not coupon.is_active
    coupon.save(update_fields=['is_active', 'updated_at'])
    return JsonResponse({'success': True, 'is_active': coupon.is_active, 'status': coupon.status})
 

@staff_member_required
@require_POST
def coupon_delete(request, pk):
    get_object_or_404(Coupon, pk=pk).delete()
    return JsonResponse({'success': True})









@staff_member_required
def review_list(request):

    status_filter = request.GET.get('status', 'all')
    search_query  = request.GET.get('q', '').strip()

    reviews = (
        Review.objects
        .select_related('product', 'product__category', 'user')
        .prefetch_related('product__variants__images')
        .order_by('-created_at')
    )

 
    if status_filter == 'pending':
        reviews = reviews.filter(is_approved=False, is_rejected=False)
    elif status_filter == 'approved':
        reviews = reviews.filter(is_approved=True)
    elif status_filter == 'rejected':
        reviews = reviews.filter(is_rejected=True)

    if search_query:
        reviews = reviews.filter(
            Q(product__name__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(comment__icontains=search_query)
        )

  
    all_reviews   = Review.objects.all()
    pending_count  = all_reviews.filter(is_approved=False, is_rejected=False).count()
    approved_count = all_reviews.filter(is_approved=True).count()
    rejected_count = all_reviews.filter(is_rejected=True).count()

    paginator = Paginator(reviews, 15)
    page      = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'adminpanel/review-list.html', {
        'reviews':        page,
        'status_filter':  status_filter,
        'search_query':   search_query,
        'pending_count':  pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
    })




@staff_member_required
@require_POST
def review_delete(request, review_id):

    review = get_object_or_404(Review, id=review_id)
    review.delete()
    return JsonResponse({'success': True, 'message': 'Review deleted.'})





def admin_logout(request):
    logout(request)
    return redirect('admin_login')












# ═══════════════════════════════════════════════════════════════════
#  ADD THESE SECTIONS TO YOUR EXISTING adminpanel/views.py
#  (paste after your existing imports and before admin_logout)
# ═══════════════════════════════════════════════════════════════════

# Add these imports at the top of adminpanel/views.py:
#
# from products.models import ProductOffer
# from orders.models import Wallet, WalletTransaction, CouponUsage
# import openpyxl
# from openpyxl.styles import Font, PatternFill, Alignment
# from io import BytesIO
# from datetime import datetime, date
# from django.db.models.functions import TruncDate, TruncWeek, TruncMonth


# ─────────────────────────────────────────────────────────────────
#  OFFER MODULE
# ─────────────────────────────────────────────────────────────────

from products.models import ProductOffer
from orders.models import Wallet, WalletTransaction, CouponUsage


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def admin_offer_list(request):
    """List all product and category offers."""
    product_offers = ProductOffer.objects.select_related('product').order_by('-created_at')
    categories     = Category.objects.filter(is_deleted=False, is_active=True).order_by('name')
    products       = Product.objects.filter(is_active=True).order_by('name')

    q = request.GET.get('q', '').strip()
    if q:
        product_offers = product_offers.filter(product__name__icontains=q)

    now = timezone.now()

    total_offers   = product_offers.count()
    active_offers  = product_offers.filter(is_active=True).count()
    expired_offers = product_offers.filter(valid_to__lt=now).count()
    cats_with_offer = categories.filter(offer_percentage__gt=0).count()

    paginator = Paginator(product_offers, 15)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'adminpanel/offer-list.html', {
        'page_obj':        page_obj,
        'categories':      categories,
        'products':        products,
        'q':               q,
        'total_offers':    total_offers,
        'active_offers':   active_offers,
        'expired_offers':  expired_offers,
        'cats_with_offer': cats_with_offer,
    })


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def admin_offer_create(request):
    """Create a product-level offer."""
    product_id          = request.POST.get('product_id')
    discount_percentage = request.POST.get('discount_percentage', '').strip()
    valid_from          = request.POST.get('valid_from', '').strip() or None
    valid_to            = request.POST.get('valid_to', '').strip()   or None
    is_active           = request.POST.get('is_active') == 'on'

    errors = []
    if not product_id:
        errors.append('Please select a product.')
    if not discount_percentage:
        errors.append('Discount percentage is required.')
    else:
        try:
            pct = int(discount_percentage)
            if not (1 <= pct <= 99):
                errors.append('Discount must be between 1 and 99.')
        except ValueError:
            errors.append('Discount must be a whole number.')

    if errors:
        return JsonResponse({'success': False, 'errors': errors})

    product = get_object_or_404(Product, pk=product_id)

    # Upsert — one offer per product (OneToOne)
    offer, created = ProductOffer.objects.update_or_create(
        product=product,
        defaults={
            'discount_percentage': int(discount_percentage),
            'valid_from':          valid_from,
            'valid_to':            valid_to,
            'is_active':           is_active,
        }
    )
    verb = 'created' if created else 'updated'
    return JsonResponse({'success': True, 'message': f'Offer {verb} for "{product.name}".'})


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def admin_offer_toggle(request, pk):
    offer            = get_object_or_404(ProductOffer, pk=pk)
    offer.is_active  = not offer.is_active
    offer.save(update_fields=['is_active', 'updated_at'])
    return JsonResponse({'success': True, 'is_active': offer.is_active})


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def admin_offer_delete(request, pk):
    get_object_or_404(ProductOffer, pk=pk).delete()
    return JsonResponse({'success': True})


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def admin_category_offer_update(request, category_id):
    """Set/update offer percentage directly on a category."""
    category = get_object_or_404(Category, pk=category_id)
    pct_raw  = request.POST.get('offer_percentage', '').strip()

    try:
        pct = int(pct_raw)
        if not (0 <= pct <= 99):
            return JsonResponse({'success': False, 'errors': ['Must be 0–99.']})
    except ValueError:
        return JsonResponse({'success': False, 'errors': ['Enter a whole number.']})

    category.offer_percentage = pct
    category.save(update_fields=['offer_percentage'])
    return JsonResponse({
        'success': True,
        'message': f'Category "{category.name}" offer set to {pct}%.',
        'offer_percentage': pct,
    })


# ─────────────────────────────────────────────────────────────────
#  RETURN MANAGEMENT — approve/reject with wallet refund
# ─────────────────────────────────────────────────────────────────

from orders.models import Order as _Order, Wallet, WalletTransaction


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def admin_approve_return(request, order_id):
    """
    Admin approves a return:
    - Items status → 'returned'
    - Stock restored
    - Wallet refund if order was prepaid
    """
    from orders.models import Wallet as _Wallet
    order = get_object_or_404(_Order, order_id=order_id)

    if order.status != 'return_requested':
        return JsonResponse({'error': 'Order is not in return_requested state.'}, status=400)

    for item in order.items.filter(status='return_requested'):
        if item.variant:
            item.variant.stock += item.quantity
            item.variant.save(update_fields=['stock'])
        item.status = 'returned'
        item.save()

    order.status = 'returned'
    order.save(update_fields=['status', 'updated_at'])

    # Wallet refund for prepaid orders
    refund_amount = order.refund_amount()
    print(f"RETURN REFUND: order={order.order_id}, method={order.payment_method}, total={order.total}, refund={refund_amount}")
    if refund_amount > 0:
        wallet, _ = _Wallet.objects.get_or_create(user=order.user)
        wallet.credit(
            refund_amount,
            description=f"Refund for returned order {order.order_id}",
            order=order,
        )
        return JsonResponse({
            'success':      True,
            'message':      f'Return approved. ₹{refund_amount:.0f} refunded to customer wallet.',
            'new_status':   'returned',
            'refund_amount': refund_amount,
        })

    return JsonResponse({
        'success':    True,
        'message':    'Return approved. COD order — no wallet refund.',
        'new_status': 'returned',
    })


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
@require_POST
def admin_reject_return(request, order_id):
    """Admin rejects a return — order goes back to delivered."""
    order = get_object_or_404(_Order, order_id=order_id)

    if order.status != 'return_requested':
        return JsonResponse({'error': 'Order is not in return_requested state.'}, status=400)

    for item in order.items.filter(status='return_requested'):
        item.status = 'active'
        item.return_reason = ''
        item.save()

    order.status = 'delivered'
    order.save(update_fields=['status', 'updated_at'])

    return JsonResponse({
        'success':    True,
        'message':    'Return request rejected. Order restored to delivered.',
        'new_status': 'delivered',
    })


# ─────────────────────────────────────────────────────────────────
#  SALES REPORT
# ─────────────────────────────────────────────────────────────────

from datetime import datetime as _datetime, timedelta as _timedelta
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate, TruncMonth
from orders.models import Order as _Order2


@never_cache
@login_required
@user_passes_test(is_admin, login_url='admin_login')
def admin_sales_report(request):
    """
    Sales report with filters: daily / weekly / monthly / custom.
    Supports PDF and Excel download via ?export=pdf|excel.
    """
    # ── date range ──
    period     = request.GET.get('period', 'monthly')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')
    export     = request.GET.get('export', '')

    today = timezone.now().date()

    if period == 'daily':
        start = today
        end   = today
    elif period == 'weekly':
        start = today - _timedelta(days=6)
        end   = today
    elif period == 'monthly':
        start = today.replace(day=1)
        end   = today
    elif period == 'yearly':
        start = today.replace(month=1, day=1)
        end   = today
    elif period == 'custom' and date_from and date_to:
        try:
            start = _datetime.strptime(date_from, '%Y-%m-%d').date()
            end   = _datetime.strptime(date_to,   '%Y-%m-%d').date()
        except ValueError:
            start, end = today.replace(day=1), today
    else:
        start = today.replace(day=1)
        end   = today

    # ── base queryset ──
    orders = _Order2.objects.filter(
        created_at__date__gte=start,
        created_at__date__lte=end,
    ).exclude(status='cancelled')

    # ── summary stats ──
    summary = orders.aggregate(
        total_orders   = Count('id'),
        gross_revenue  = Sum('subtotal'),
        total_discount = Sum('discount'),
        total_shipping = Sum('shipping'),
        net_revenue    = Sum('total'),
    )
    for key in summary:
        if summary[key] is None:
            summary[key] = 0

    # ── daily breakdown ──
    daily_data = (
        orders
        .annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(
            orders_count=Count('id'),
            gross       =Sum('subtotal'),
            discount    =Sum('discount'),
            net         =Sum('total'),
        )
        .order_by('day')
    )

    # ── top products ──
    from orders.models import OrderItem as _OI
    top_products = (
        _OI.objects.filter(
            order__created_at__date__gte=start,
            order__created_at__date__lte=end,
        )
        .exclude(order__status='cancelled')
        .values('product_name')
        .annotate(units=Sum('quantity'), revenue=Sum('unit_price'))
        .order_by('-units')[:10]
    )

    # ── payment method breakdown ──
    payment_breakdown = (
        orders.values('payment_method')
        .annotate(count=Count('id'), total=Sum('total'))
        .order_by('-total')
    )

    context = {
        'period':            period,
        'start':             start,
        'end':               end,
        'date_from':         date_from,
        'date_to':           date_to,
        'summary':           summary,
        'daily_data':        list(daily_data),
        'top_products':      list(top_products),
        'payment_breakdown': list(payment_breakdown),
    }

    # ── exports ──
    if export == 'pdf':
        return _export_sales_pdf(context)
    if export == 'excel':
        return _export_sales_excel(context)

    return render(request, 'adminpanel/sales-report.html', context)


def _export_sales_pdf(ctx):
    """Generate and return PDF sales report."""
    try:
        from io import BytesIO
        from reportlab.lib              import colors
        from reportlab.lib.pagesizes    import A4, landscape
        from reportlab.lib.styles       import ParagraphStyle
        from reportlab.lib.units        import mm
        from reportlab.platypus         import (
            HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        )

        buffer  = BytesIO()
        doc     = SimpleDocTemplate(
            buffer, pagesize=landscape(A4),
            rightMargin=15*mm, leftMargin=15*mm,
            topMargin=15*mm,   bottomMargin=15*mm
        )
        h1  = ParagraphStyle('H1',  fontSize=20, fontName='Helvetica-Bold',
                              textColor=colors.HexColor('#2a2520'), spaceAfter=4)
        h2  = ParagraphStyle('H2',  fontSize=12, fontName='Helvetica-Bold',
                              textColor=colors.HexColor('#2a2520'), spaceBefore=10, spaceAfter=4)
        sub = ParagraphStyle('Sub', fontSize=9,  fontName='Helvetica',
                              textColor=colors.HexColor('#888888'), spaceAfter=8)

        s   = ctx['summary']
        story = []
        story.append(Paragraph('ALAIA — Sales Report', h1))
        story.append(Paragraph(
            f"Period: {ctx['start'].strftime('%d %b %Y')} → {ctx['end'].strftime('%d %b %Y')}",
            sub
        ))
        story.append(HRFlowable(width='100%', thickness=1,
                                 color=colors.HexColor('#c9a96e'), spaceAfter=8))

        # Summary table
        story.append(Paragraph('Summary', h2))
        sum_data = [
            ['Total Orders', 'Gross Revenue', 'Total Discount', 'Shipping', 'Net Revenue'],
            [
                str(s['total_orders']),
                f"₹{s['gross_revenue']:,.0f}",
                f"₹{s['total_discount']:,.0f}",
                f"₹{s['total_shipping']:,.0f}",
                f"₹{s['net_revenue']:,.0f}",
            ]
        ]
        sum_tbl = Table(sum_data, colWidths=[40*mm]*5)
        sum_tbl.setStyle(TableStyle([
            ('BACKGROUND',   (0,0), (-1,0), colors.HexColor('#2a2520')),
            ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
            ('FONTNAME',     (0,0), (-1,-1),'Helvetica-Bold'),
            ('FONTSIZE',     (0,0), (-1,-1), 9),
            ('ALIGN',        (0,0), (-1,-1),'CENTER'),
            ('GRID',         (0,0), (-1,-1), 0.5, colors.HexColor('#dddddd')),
            ('BOTTOMPADDING',(0,0), (-1,-1), 6),
            ('TOPPADDING',   (0,0), (-1,-1), 6),
        ]))
        story.append(sum_tbl)
        story.append(Spacer(1, 6*mm))

        # Daily breakdown
        if ctx['daily_data']:
            story.append(Paragraph('Daily Breakdown', h2))
            rows = [['Date', 'Orders', 'Gross (₹)', 'Discount (₹)', 'Net (₹)']]
            for row in ctx['daily_data']:
                rows.append([
                    row['day'].strftime('%d %b %Y'),
                    str(row['orders_count']),
                    f"{row['gross'] or 0:,.0f}",
                    f"{row['discount'] or 0:,.0f}",
                    f"{row['net'] or 0:,.0f}",
                ])
            tbl = Table(rows, colWidths=[40*mm, 25*mm, 40*mm, 40*mm, 40*mm])
            tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,0), colors.HexColor('#555555')),
                ('TEXTCOLOR',     (0,0), (-1,0), colors.white),
                ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTNAME',      (0,1), (-1,-1),'Helvetica'),
                ('FONTSIZE',      (0,0), (-1,-1), 8),
                ('ALIGN',         (1,0), (-1,-1),'RIGHT'),
                ('ROWBACKGROUNDS',(0,1), (-1,-1),[colors.white, colors.HexColor('#f9f9f9')]),
                ('GRID',          (0,0), (-1,-1), 0.3, colors.HexColor('#dddddd')),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            story.append(tbl)

        doc.build(story)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="ALAIA-Sales-Report-{ctx["start"]}-{ctx["end"]}.pdf"'
        )
        return response

    except ImportError:
        return HttpResponse("pip install reportlab", content_type='text/plain')


def _export_sales_excel(ctx):
    """Generate and return Excel sales report."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = 'Sales Report'

        GOLD  = 'C9A96E'
        DARK  = '2A2520'
        LIGHT = 'FDF9F4'

        header_font  = Font(bold=True, color='FFFFFF', size=11)
        header_fill  = PatternFill('solid', fgColor=DARK)
        center       = Alignment(horizontal='center', vertical='center')
        thin_border  = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        def hrow(ws, row, values, fill_color=DARK):
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = Font(bold=True, color='FFFFFF', size=10)
                c.fill      = PatternFill('solid', fgColor=fill_color)
                c.alignment = center
                c.border    = thin_border

        def drow(ws, row, values):
            for col, val in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.alignment = Alignment(horizontal='right' if col > 1 else 'left')
                c.border    = thin_border
                if row % 2 == 0:
                    c.fill = PatternFill('solid', fgColor='F5F0EA')

        # Title
        ws.merge_cells('A1:E1')
        title = ws['A1']
        title.value     = f"ALAIA Sales Report | {ctx['start']} → {ctx['end']}"
        title.font      = Font(bold=True, size=14, color=DARK)
        title.alignment = center
        title.fill      = PatternFill('solid', fgColor='FDF9F4')
        ws.row_dimensions[1].height = 30

        # Summary
        ws.append([])
        hrow(ws, 3, ['Total Orders', 'Gross Revenue', 'Total Discount', 'Shipping', 'Net Revenue'])
        s = ctx['summary']
        drow(ws, 4, [
            s['total_orders'],
            f"₹{s['gross_revenue']:,.0f}",
            f"₹{s['total_discount']:,.0f}",
            f"₹{s['total_shipping']:,.0f}",
            f"₹{s['net_revenue']:,.0f}",
        ])

        # Daily breakdown
        ws.append([])
        ws.append([])
        r = ws.max_row
        hrow(ws, r, ['Date', 'Orders', 'Gross (₹)', 'Discount (₹)', 'Net (₹)'])
        for row in ctx['daily_data']:
            r += 1
            drow(ws, r, [
                row['day'].strftime('%d %b %Y'),
                row['orders_count'],
                row['gross']    or 0,
                row['discount'] or 0,
                row['net']      or 0,
            ])

        # Top products
        if ctx['top_products']:
            ws.append([])
            ws.append([])
            r = ws.max_row
            hrow(ws, r, ['Product', 'Units Sold', 'Revenue (₹)', '', ''])
            for row in ctx['top_products']:
                r += 1
                drow(ws, r, [row['product_name'], row['units'], row['revenue'] or 0, '', ''])

        # Column widths
        for col in ws.columns:
            max_len = 10
            col_letter = None
            for c in col:
                if hasattr(c, 'column_letter'):
                    col_letter = c.column_letter
                    if c.value:
                        max_len = max(max_len, len(str(c.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = (
            f'attachment; filename="ALAIA-Sales-{ctx["start"]}-{ctx["end"]}.xlsx"'
        )
        return response

    except ImportError:
        return HttpResponse("pip install openpyxl", content_type='text/plain')